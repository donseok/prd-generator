#!/usr/bin/env python3
"""WBS Excel maker script.

WBS JSON을 읽어 업로드 양식에 맞는 Excel(XLSX) 파일을 생성합니다.

Usage:
    python -m app.scripts.wbs_excel_maker
"""

import json
import math
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

# --- 스타일 상수 ---
FONT_NAME = "Pretendard"
FONT_FALLBACK = "맑은 고딕"

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

# Row 1: 설명 헤더
DESC_FONT = Font(name=FONT_NAME, size=8, color="64748B")
DESC_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
DESC_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row 2: 컬럼명 헤더
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Phase 행
PHASE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
PHASE_NAME_FONT = Font(name=FONT_NAME, size=10, bold=True, color="1E40AF")
PHASE_FONT = Font(name=FONT_NAME, size=10, color="475569")

# Activity 행
ACTIVITY_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
ACTIVITY_NAME_FONT = Font(name=FONT_NAME, size=10, bold=True, color="3730A3")
ACTIVITY_FONT = Font(name=FONT_NAME, size=10, color="475569")

# Task 행
TASK_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
TASK_NAME_FONT = Font(name=FONT_NAME, size=10, color="334155")
TASK_FONT = Font(name=FONT_NAME, size=10, color="475569")

# 컬럼 정의
COLUMNS = {
    "A": {"width": 12, "desc": "계층 번호\n(예: 1, 1.1, 1.1.1)", "header": "WBS코드"},
    "B": {"width": 14, "desc": "상위 작업의\nWBS코드", "header": "상위WBS코드"},
    "C": {"width": 12, "desc": "Phase / Activity\n/ Task", "header": "구분"},
    "D": {"width": 40, "desc": "작업 이름\n(필수)", "header": "작업명"},
    "E": {"width": 20, "desc": "산출물명\n(선택)", "header": "산출물"},
    "F": {"width": 12, "desc": "담당자 이름\n(선택)", "header": "담당자"},
    "G": {"width": 10, "desc": "리프 작업의\n비중 (합계\u2248100)", "header": "가중치"},
    "H": {"width": 13, "desc": "예상 소요일\n(리프 작업만)", "header": "기간일수"},
    "I": {"width": 14, "desc": "YYYY-MM-DD\n형식", "header": "계획시작"},
    "J": {"width": 13, "desc": "YYYY-MM-DD\n형식", "header": "계획종료"},
    "K": {"width": 10, "desc": "대기 / 진행중\n/ 완료 / 보류", "header": "상태"},
}

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(vertical="center")

STATUS_MAP = {
    "NOT_STARTED": "대기",
    "IN_PROGRESS": "진행중",
    "COMPLETED": "완료",
    "ON_HOLD": "보류",
}


def find_latest_wbs_json() -> Path | None:
    """workspace/outputs/wbs/ 에서 최신 WBS JSON 파일을 찾습니다."""
    wbs_dir = Path("workspace/outputs/wbs")
    json_files = list(wbs_dir.glob("WBS-*.json"))
    if not json_files:
        return None
    return max(json_files, key=lambda x: x.stat().st_mtime)


def load_wbs_data(path: Path) -> dict:
    """WBS JSON 파일을 로드합니다."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def flatten_wbs_to_rows(wbs_data: dict) -> list[dict]:
    """WBS JSON 구조를 Excel 행 리스트로 변환합니다.

    3단계 계층: Phase(1) → Activity(1.1) → Task(1.1.1)
    """
    rows = []
    total_task_hours = 0

    # 먼저 전체 task 시간 합산 (가중치 계산용)
    for phase in wbs_data.get("phases", []):
        for wp in phase.get("work_packages", []):
            for task in wp.get("tasks", []):
                total_task_hours += task.get("estimated_hours", 0)

    phase_num = 0
    for phase in wbs_data.get("phases", []):
        phase_num += 1
        phase_code = str(phase_num)

        # Phase 행
        rows.append({
            "wbs_code": phase_code,
            "parent_code": "",
            "level": "Phase",
            "name": phase.get("name", ""),
            "deliverable": "",
            "assignee": "",
            "weight": "",
            "duration_days": "",
            "plan_start": "",
            "plan_end": "",
            "status": "",
        })

        wp_num = 0
        for wp in phase.get("work_packages", []):
            wp_num += 1
            wp_code = f"{phase_code}.{wp_num}"

            # Activity 행
            rows.append({
                "wbs_code": wp_code,
                "parent_code": phase_code,
                "level": "Activity",
                "name": wp.get("name", ""),
                "deliverable": "",
                "assignee": "",
                "weight": "",
                "duration_days": "",
                "plan_start": "",
                "plan_end": "",
                "status": "",
            })

            task_num = 0
            for task in wp.get("tasks", []):
                task_num += 1
                task_code = f"{wp_code}.{task_num}"
                hours = task.get("estimated_hours", 0)

                # 가중치 계산
                weight = ""
                if total_task_hours > 0 and hours > 0:
                    weight = round(hours / total_task_hours * 100, 1)

                # 기간일수
                duration = ""
                if hours > 0:
                    duration = max(1, math.ceil(hours / 8))

                # 담당자
                assignee = ""
                resources = task.get("resources", [])
                if resources:
                    assignee = resources[0].get("resource_type", "")

                # 산출물
                deliverables = task.get("deliverables", [])
                deliverable = ", ".join(deliverables) if deliverables else ""

                # 날짜
                start_date = task.get("start_date", "") or ""
                end_date = task.get("end_date", "") or ""

                # 상태
                status = STATUS_MAP.get(task.get("status", "NOT_STARTED"), "대기")

                rows.append({
                    "wbs_code": task_code,
                    "parent_code": wp_code,
                    "level": "Task",
                    "name": task.get("name", ""),
                    "deliverable": deliverable,
                    "assignee": assignee,
                    "weight": weight,
                    "duration_days": duration,
                    "plan_start": start_date,
                    "plan_end": end_date,
                    "status": status,
                })

    return rows


def create_data_sheet(wb: Workbook, rows: list[dict]):
    """'WBS 데이터' 시트를 생성합니다."""
    ws = wb.active
    ws.title = "WBS 데이터"

    # 컬럼 너비 설정
    for col_letter, col_def in COLUMNS.items():
        ws.column_dimensions[col_letter].width = col_def["width"]

    # Row 1: 설명 헤더
    for i, (col_letter, col_def) in enumerate(COLUMNS.items(), 1):
        cell = ws.cell(row=1, column=i, value=col_def["desc"])
        cell.font = DESC_FONT
        cell.fill = DESC_FILL
        cell.alignment = DESC_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 36

    # Row 2: 컬럼명 헤더
    for i, (col_letter, col_def) in enumerate(COLUMNS.items(), 1):
        cell = ws.cell(row=2, column=i, value=col_def["header"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 30

    # 데이터 행
    for row_idx, row_data in enumerate(rows, 3):
        level = row_data["level"]

        if level == "Phase":
            fill = PHASE_FILL
            name_font = PHASE_NAME_FONT
            data_font = PHASE_FONT
        elif level == "Activity":
            fill = ACTIVITY_FILL
            name_font = ACTIVITY_NAME_FONT
            data_font = ACTIVITY_FONT
        else:
            fill = TASK_FILL
            name_font = TASK_NAME_FONT
            data_font = TASK_FONT

        values = [
            row_data["wbs_code"],
            row_data["parent_code"],
            row_data["level"],
            row_data["name"],
            row_data["deliverable"],
            row_data["assignee"],
            row_data["weight"],
            row_data["duration_days"],
            row_data["plan_start"],
            row_data["plan_end"],
            row_data["status"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = THIN_BORDER

            # D 열(작업명)은 별도 폰트, 좌측 정렬
            if col_idx == 4:
                cell.font = name_font
                cell.alignment = LEFT_ALIGN
            # E, F 열(산출물, 담당자)은 좌측 정렬
            elif col_idx in (5, 6):
                cell.font = data_font
                cell.alignment = LEFT_ALIGN
            else:
                cell.font = data_font
                cell.alignment = CENTER_ALIGN

        ws.row_dimensions[row_idx].height = 24

    # 빈 값("") 처리 - 빈 문자열 대신 None으로
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=11):
        for cell in row:
            if cell.value == "":
                cell.value = None


def create_guide_sheet(wb: Workbook):
    """'작성 가이드' 시트를 생성합니다."""
    ws = wb.create_sheet(title="작성 가이드")

    # 컬럼 너비
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70

    title_font = Font(name=FONT_NAME, size=14, bold=True, color="1E293B")
    section_font = Font(name=FONT_NAME, size=11, bold=True, color="1E40AF")
    label_font = Font(name=FONT_NAME, size=10, bold=True, color="334155")
    desc_font = Font(name=FONT_NAME, size=10, color="475569")
    section_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

    guide_data = [
        ("title", "WBS 업로드 양식 — 작성 가이드", None),
        ("empty", None, None),
        ("section", "1. 기본 규칙", None),
        ("item", "WBS코드", "계층 번호 체계입니다. 점(.)으로 레벨을 구분합니다.\n  Phase: 1, 2, 3 ...\n  Activity: 1.1, 1.2, 2.1 ...\n  Task: 1.1.1, 1.1.2, 2.1.1 ..."),
        ("item", "상위WBS코드", "WBS코드에서 마지막 번호를 뺀 상위 코드입니다.\n  1.1의 상위 → 1\n  1.1.2의 상위 → 1.1\n  Phase(최상위)는 빈칸으로 둡니다."),
        ("item", "구분", "Phase(단계), Activity(활동), Task(작업) 중 하나를 선택합니다."),
        ("item", "작업명", "작업의 이름입니다. 필수 입력 항목입니다."),
        ("empty", None, None),
        ("section", "2. 계층 구조", None),
        ("item", "Phase (1레벨)", "프로젝트의 대단계입니다. (예: 분석, 설계, 개발, 테스트)\n날짜/공정율/상태는 하위 작업에서 자동 계산되므로 비워두세요."),
        ("item", "Activity (2레벨)", "Phase 하위의 주요 활동입니다.\n날짜/공정율/상태는 하위 작업에서 자동 계산되므로 비워두세요."),
        ("item", "Task (3레벨)", "실제 수행하는 작업(리프 태스크)입니다.\n담당자, 가중치, 기간일수, 계획시작/종료, 상태를 입력하세요."),
        ("empty", None, None),
        ("section", "3. 필드 설명", None),
        ("item", "산출물", "해당 작업의 결과물 이름입니다. (선택)"),
        ("item", "담당자", "작업 담당자의 이름입니다. (선택)"),
        ("item", "가중치", "리프(Task) 작업의 비중입니다. 모든 리프 작업의 가중치 합계가\n100에 가깝도록 배분하세요. 공정율 계산에 사용됩니다."),
        ("item", "기간일수", "예상 소요일입니다. (선택) 리프 작업에만 입력합니다."),
        ("item", "계획시작 / 종료", "날짜 형식: YYYY-MM-DD (예: 2026-04-01)"),
        ("item", "상태", "대기, 진행중, 완료, 보류 중 선택합니다.\n비워두면 \"대기\"로 설정됩니다."),
        ("empty", None, None),
        ("section", "4. 주의사항", None),
        ("item", "자동 계산", "Phase/Activity의 계획시작, 계획종료, 공정율, 상태는\n하위 Task들로부터 자동으로 집계됩니다."),
        ("item", "순서", "같은 레벨 내에서 행 순서대로 정렬됩니다.\nWBS코드 번호 순서와 행 순서를 일치시키세요."),
        ("item", "시트 이름", "\"WBS 데이터\" 시트 이름을 변경하지 마세요.\n시스템이 이 시트 이름으로 데이터를 찾습니다."),
    ]

    row = 0
    for entry_type, col_b, col_c in guide_data:
        row += 1
        if entry_type == "title":
            cell = ws.cell(row=row, column=2, value=col_b)
            cell.font = title_font
            ws.merge_cells(f"B{row}:C{row}")
        elif entry_type == "section":
            cell = ws.cell(row=row, column=2, value=col_b)
            cell.font = section_font
            ws.merge_cells(f"B{row}:C{row}")
            for c in range(2, 4):
                ws.cell(row=row, column=c).fill = section_fill
        elif entry_type == "item":
            cell_b = ws.cell(row=row, column=2, value=col_b)
            cell_b.font = label_font
            cell_b.alignment = Alignment(vertical="top")
            cell_c = ws.cell(row=row, column=3, value=col_c)
            cell_c.font = desc_font
            cell_c.alignment = Alignment(vertical="top", wrap_text=True)


def generate_wbs_excel(wbs_data: dict, output_path: Path):
    """WBS 데이터를 Excel 파일로 생성합니다."""
    wb = Workbook()

    # WBS 데이터를 행으로 변환
    rows = flatten_wbs_to_rows(wbs_data)

    # 시트 생성
    create_data_sheet(wb, rows)
    create_guide_sheet(wb)

    # 저장
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return len(rows)


def main():
    print("\n" + "=" * 70)
    print("WBS Excel 생성 시작")
    print(f'시작 시간: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print("=" * 70)

    # 최신 WBS JSON 찾기
    wbs_path = find_latest_wbs_json()
    if not wbs_path:
        print("\nWBS JSON 파일을 찾을 수 없습니다.")
        print("먼저 /wbs:wbs-maker를 실행하세요.")
        return

    print(f"\n입력 WBS: {wbs_path}")

    # WBS 로드
    wbs_data = load_wbs_data(wbs_path)
    title = wbs_data.get("title", "WBS")
    print(f"WBS 제목: {title}")

    phase_count = len(wbs_data.get("phases", []))
    task_count = sum(
        len(task)
        for phase in wbs_data.get("phases", [])
        for task in [
            t for wp in phase.get("work_packages", []) for t in [wp.get("tasks", [])]
        ]
    )
    print(f"총 단계: {phase_count}개")

    # Excel 생성
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_dir = Path("workspace/outputs/wbs")
    output_path = output_dir / f"WBS_업로드_{timestamp}.xlsx"

    row_count = generate_wbs_excel(wbs_data, output_path)

    print("\n" + "=" * 70)
    print("WBS Excel 생성 완료")
    print("=" * 70)
    print(f"  총 행 수: {row_count}개 (Phase + Activity + Task)")
    print(f"  출력 파일: {output_path}")
    print(f'  완료 시간: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')


if __name__ == "__main__":
    main()
